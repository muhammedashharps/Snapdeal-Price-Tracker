from bs4 import BeautifulSoup as bs
import requests
import yagmail
from openpyxl import load_workbook

wb = load_workbook("snapdeal_excel.xlsx")
ws = wb.active
row_count = ws.max_row

print("Processing...........")


def send_alert(reciever):
    for col in range(2, row_count + 1):
        link = ws[f"B{col}"].value
        pricing = ws[f"C{col}"].value
        name = ws[f"A{col}"].value.lstrip()

        response = requests.get(link, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9"})

        html = response.text
        data = bs(html, "html.parser")
        name = data.find("h1", itemprop="name").get_text().lstrip()
        latest_price = float(data.find("span", itemprop="price").get_text().replace(",", ""))

        if (latest_price < pricing):
            subject = f"Price Drop Alert: {name}"
            contents = [
                f"We are pleased to inform you that the price of '{name}' on Snapdeal has dropped to a new low of Rs,{latest_price}. This price reduction is below your desired threshold of Rs{pricing}, making it an excellent time to consider your purchase.\n\n",

                f"Product Details:\n\n",
                f"Product Name: {name}\n\n",
                f"Current Price: Rs,{latest_price}\n",
                f"Original Price: Rs,{pricing}\n\n",
                f"Product Link: {link}\n\n",

                f"Hurry and seize this opportunity to make your purchase while the price is favorable.\n\n",

                f"Thank you for using our SnapDeal price tracking service",

            ]
            contents = ''.join(contents)

            yagmail.SMTP('Your Email', 'Your Password').send(reciever, subject, contents)
            print(f"Price Drop Detected! . Succesfully Send The Alert Message To {reciever}")
        ws[f"C{col}"].value = latest_price


if __name__ == "__main__":

    reciever = "ashharjosh@gmail.com"

    try:
        send_alert(reciever)
    except:
        print("Error Tracking The Prices")

    wb.save("snapdeal_excel.xlsx")
    print("Finished Checking For Price Alert")
