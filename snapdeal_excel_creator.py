import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup as bs

wb = load_workbook('snapdeal_excel.xlsx')
ws = wb.active

names = []
links = []
prices = []

selection = int(input("""Enter the number corresponding to the option you are selecting
                      1. Add new product to the tracking list
                      2. See all the products and remove the product that is no longer required
                       : """))

if selection == 1:
    def extract_data(products):
        for i in range(1, products + 1):
            link = input(f"Enter The Link Of The Product {i}: ")
            response = requests.get(link, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                                                 "AppleWebKit/537.36 (KHTML, like Gecko) "
                                                                 "Chrome/116.0.0.0 Safari/537.36",
                                                   "Accept-Language": "en-US,en;q=0.9"})
            html = response.text
            data = bs(html, "html.parser")
            name = data.find("h1", itemprop="name").get_text().lstrip()
            price = float(data.find("span", itemprop="price").get_text().replace(",", ""))
            links.append(link)
            prices.append(price)
            names.append(name)


    def save():
        for i in range(products):
            ws.append([names[i], links[i], prices[i]])
            wb.save("snapdeal_excel.xlsx")
        print("Products Successfully Added To The Excel Sheet")


    while True:
        try:
            products = int(input("How many products do you want to add to track: "))
            break
        except ValueError:
            print("Error! Please enter an integer value.")
    try:
        extract_data(products)
        save()
    except:
        print("Error Extracting The Data")

if selection == 2:
    def print_products():
        row_count = ws.max_row
        print("Following are the Products to Be Tracked: ")
        for col in range(2, row_count + 1):
            name = ws[f"A{col}"].value.lstrip()
            print(f"{col}.{name} ")


    print_products()

    def remove():
        while True:
            remove_a_product = input("Do You Want To Remove any Product (Y/N) : ").upper()

            if remove_a_product == "N":
                print("No Product Removed")
                break

            elif remove_a_product == "Y":
                number = int(input("Enter the Sl No. of the product you want to remove from the list (Row Number) : "))
                ws.delete_rows(number)
                print("Product Successfully Removed")
                wb.save("snapdeal_excel.xlsx")
                break

            else:
                print("Error! Enter Either Y or N")

    try:
        remove()
    except:
        print("Failed To Remove The Product !!!, ReCheck your entered values")
